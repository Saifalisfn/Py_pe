#!/usr/bin/env python3
"""
Excel Export Utility
Usage:
  python export.py transactions          # export all transactions
  python export.py transactions 8477054372   # export for one mobile
  python export.py utr_log               # export UTR search history
"""
import sys
import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import db
from rich.console import Console

console = Console()

# ─── Styling helpers ─────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="F2F4F8")
BORDER      = Border(
    bottom=Side(style="thin", color="D0D4DC"),
    top=Side(style="thin", color="D0D4DC"),
)

def _style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

def _style_rows(ws, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start), 0):
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")


# ─── Exporters ───────────────────────────────────────────────────────────────

def export_transactions(mobile=None, filename=None):
    rows = db.get_transactions_by_mobile(mobile, limit=10000) if mobile else \
           db.fetchall("SELECT * FROM transactions ORDER BY transaction_date DESC LIMIT 10000")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.row_dimensions[1].height = 28

    headers = ["#", "Date", "Account", "Amount (₹)", "Status", "Payer", "Payer VPA", "UTR", "Transaction ID"]
    _style_header(ws, headers)

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            str(r.get("transaction_date", "") or ""),
            r.get("session_mobile", ""),
            float(r.get("amount", 0) or 0),
            r.get("status", ""),
            r.get("payer_name", "") or "",
            r.get("payer_vpa", "") or "",
            r.get("utr", "") or "",
            r.get("transaction_id", "") or "",
        ])

    _style_rows(ws)
    _autofit(ws)

    if not filename:
        tag = f"_{mobile}" if mobile else ""
        filename = f"transactions{tag}_{_ts()}.xlsx"

    wb.save(filename)
    return os.path.abspath(filename)


def export_utr_results(results, filename=None):
    """Export batch UTR search results (list of dicts from batch_utr.py)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UTR Results"
    ws.row_dimensions[1].height = 28

    headers = ["UTR", "Found", "Amount (₹)", "Status", "Payer", "Account"]
    _style_header(ws, headers)

    GREEN = PatternFill("solid", fgColor="D4EDDA")
    RED   = PatternFill("solid", fgColor="F8D7DA")

    for i, r in enumerate(results, 2):
        ws.append([
            r["utr"],
            "Yes" if r["found"] else "No",
            float(r["amount"]) if r["amount"] else "",
            r["status"] or "",
            r["payer"] or "",
            r["account"] or "",
        ])
        fill = GREEN if r["found"] else RED
        for cell in ws[i]:
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    _autofit(ws)

    if not filename:
        filename = f"utr_batch_{_ts()}.xlsx"

    wb.save(filename)
    return os.path.abspath(filename)


def export_utr_log(filename=None):
    rows = db.get_utr_search_history(limit=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UTR Search Log"
    ws.row_dimensions[1].height = 28

    headers = ["#", "UTR", "Found", "Amount (₹)", "Status", "Account", "Searched At"]
    _style_header(ws, headers)

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r.get("utr", ""),
            "Yes" if r.get("found") else "No",
            float(r.get("amount", 0) or 0) or "",
            r.get("status", "") or "",
            r.get("session_mobile", "") or "",
            str(r.get("searched_at", "") or ""),
        ])

    _style_rows(ws)
    _autofit(ws)

    if not filename:
        filename = f"utr_log_{_ts()}.xlsx"

    wb.save(filename)
    return os.path.abspath(filename)


def _ts():
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    args = sys.argv[1:]
    mode = args[0] if args else ""

    if mode == "transactions":
        mobile = args[1] if len(args) > 1 else None
        path = export_transactions(mobile)
        console.print(f"[green]Exported transactions → {path}[/green]")

    elif mode == "utr_log":
        path = export_utr_log()
        console.print(f"[green]Exported UTR log → {path}[/green]")

    else:
        console.print(
            "[bold cyan]Export Utility[/bold cyan]\n\n"
            "  [white]python export.py transactions[/white]            — all transactions\n"
            "  [white]python export.py transactions 8477054372[/white] — one account\n"
            "  [white]python export.py utr_log[/white]                 — UTR search history\n"
        )

    db.close()
