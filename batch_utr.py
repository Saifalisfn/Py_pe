#!/usr/bin/env python3
"""
Batch UTR Verification
Usage:
  python batch_utr.py 123456789 987654321 ...     # pass UTRs as args
  python batch_utr.py utrs.txt                    # one UTR per line in txt file
  python batch_utr.py utrs.xlsx                   # xlsx with UTR column
  python batch_utr.py                             # interactive input
"""
import sys
import os
import time
from rich.console import Console
from rich.table import Table
from rich.progress import track
from utr_finder import find_utr
from database import db

console = Console()


def _read_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Find UTR column: header row first
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    utr_col = next((i for i, h in enumerate(headers) if "utr" in h), 0)  # default col 0

    utrs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[utr_col] if utr_col < len(row) else None
        if val:
            utrs.append(str(val).strip())

    wb.close()
    return utrs


def load_utrs(args):
    if not args:
        raw = console.input("[bold white]Enter UTR numbers (comma or space separated): [/bold white]")
        return [u.strip() for u in raw.replace(",", " ").split() if u.strip()]

    if len(args) == 1 and os.path.isfile(args[0]):
        path = args[0]
        if path.lower().endswith(".xlsx"):
            utrs = _read_xlsx(path)
        elif path.lower().endswith(".csv"):
            import csv
            with open(path, newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
            # find UTR column
            headers = [h.strip().lower() for h in rows[0]] if rows else []
            col = next((i for i, h in enumerate(headers) if "utr" in h), 0)
            start = 1 if headers else 0
            utrs = [r[col].strip() for r in rows[start:] if r and r[col].strip()]
        else:
            with open(path) as f:
                utrs = [line.strip() for line in f if line.strip()]
        console.print(f"[dim]Loaded {len(utrs)} UTRs from {path}[/dim]")
        return utrs

    return list(args)


def run_batch(utrs):
    results = []
    console.print(f"\n[bold cyan]Searching {len(utrs)} UTR(s)...[/bold cyan]\n")

    for utr in track(utrs, description="Searching..."):
        found, data = find_utr(utr)
        results.append({
            "utr": utr,
            "found": found,
            "amount": data.get("amount") if data else None,
            "status": data.get("status") if data else None,
            "payer": (data.get("payer_name") or data.get("payerName")) if data else None,
            "account": data.get("session_mobile") if data else None,
        })
        time.sleep(0.3)  # avoid rate limiting

    return results


def print_results(results):
    found = [r for r in results if r["found"]]
    not_found = [r for r in results if not r["found"]]

    table = Table(title=f"Batch UTR Results ({len(found)}/{len(results)} found)")
    table.add_column("UTR", style="white")
    table.add_column("Found", style="bold")
    table.add_column("Amount", style="green")
    table.add_column("Status", style="cyan")
    table.add_column("Payer", style="yellow")
    table.add_column("Account", style="magenta")

    for r in results:
        table.add_row(
            r["utr"],
            "[green]✓[/green]" if r["found"] else "[red]✗[/red]",
            f"₹{r['amount']}" if r["amount"] else "—",
            r["status"] or "—",
            r["payer"] or "—",
            r["account"] or "—",
        )

    console.print(table)
    console.print(f"\n[green]Found: {len(found)}[/green]  [red]Not Found: {len(not_found)}[/red]")
    return results


def save_results(results):
    answer = console.input("\n[bold white]Export to Excel? (y/n): [/bold white]").strip().lower()
    if answer == "y":
        from export import export_utr_results
        path = export_utr_results(results)
        console.print(f"[green]Saved: {path}[/green]")


if __name__ == "__main__":
    utrs = load_utrs(sys.argv[1:])
    if not utrs:
        console.print("[red]No UTRs provided.[/red]")
        sys.exit(1)

    results = run_batch(utrs)
    print_results(results)
    save_results(results)

    db.close()
